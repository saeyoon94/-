import pandas as pd
from bs4 import BeautifulSoup
import urllib
import time
from urllib.request import urlopen
import openpyxl



code = pd.read_csv('code.csv') #투자대상종목들의 종목코드
after16 = pd.read_csv('after16.csv')  #16년이후 상장된 종목들의 종목코드(16년부터 학습시키므로, 학습시에는 제외시켜야한다)

code = list(code)
after16 = list(after16)

transfer = []

for i in code :
    new = str(i[1:])
    transfer.append(new)

code = transfer

transfer = []

for i in after16 :
    new = str(i[1:])
    transfer.append(new)

after16 = transfer

for i in code :                  #code에서 2016년이후 상장된 종목들은 제거
    if i in after16 :
        code.remove(i)

excel = openpyxl.Workbook()  #엑셀 생성

#스크레이핑

for index, stockItem in enumerate(code) :              #모든 종목코드에 대해서 스크레이핑
    print(index)
    mpNum = 25 #16년 1월부터 긁어온다
    url = 'http://finance.naver.com/item/frgn.nhn?code=' + stockItem
    html = urlopen(url)
    source = BeautifulSoup(html.read(), "html.parser")
    name = source.select("div.wrap_company h2")

    sheet = excel.create_sheet(index=index, title=name[0].text)   #회사이름의 시트 생성

    for page in range(1, mpNum + 1):

        url = 'http://finance.naver.com/item/frgn.nhn?code=' + stockItem + '&page=' + str(page)
        html = urlopen(url)
        source = BeautifulSoup(html.read(), "html.parser")
        srlists = source.find_all("tr", onmouseover = 'mouseOver(this)')
        isCheckNone = None

        if ((page % 1) == 0):
            time.sleep(0.5)


        for i in range(len(srlists)):
            if (srlists[i].span != isCheckNone):
                srlists[i].td.text
                sheet.cell(row=1, column= i+1+20*(page-1), value=srlists[i].find_all("td", class_="num")[0].text)   #주가
                sheet.cell(row=2, column=i + 1 + 20 * (page - 1), value=srlists[i].find_all("td", class_="num")[3].text)  # 거래량
                sheet.cell(row=3, column=i + 1 + 20 * (page - 1), value=srlists[i].find_all("td", class_="num")[4].text)  # 기관순매수
                sheet.cell(row=4, column=i + 1 + 20 * (page - 1), value=srlists[i].find_all("td", class_="num")[5].text)  # 외국인순매수



excel.save('stock_data.xlsx')


















